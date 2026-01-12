"""
EXIF Extractors - Low-level metadata extraction from various file types.

Supports:
- Images: JPEG, PNG, TIFF, HEIC, WebP (EXIF, IPTC, XMP)
- Documents: PDF, DOCX, XLSX, PPTX
- Videos: MP4, MOV, AVI (basic metadata)
"""

import io
import logging
import subprocess
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


@dataclass
class MetadataResult:
    """Result of metadata extraction."""
    url: str
    file_type: str
    success: bool
    metadata: Dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None

    # Common investigation-relevant fields (extracted for convenience)
    author: Optional[str] = None
    creator: Optional[str] = None
    company: Optional[str] = None
    software: Optional[str] = None
    create_date: Optional[str] = None
    modify_date: Optional[str] = None
    original_date: Optional[str] = None  # DateTimeOriginal - when photo was taken
    gps_lat: Optional[float] = None
    gps_lon: Optional[float] = None
    camera_make: Optional[str] = None
    camera_model: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "url": self.url,
            "file_type": self.file_type,
            "success": self.success,
            "metadata": self.metadata,
            "error": self.error,
            "author": self.author,
            "creator": self.creator,
            "company": self.company,
            "software": self.software,
            "create_date": self.create_date,
            "modify_date": self.modify_date,
            "original_date": self.original_date,
            "gps": {"lat": self.gps_lat, "lon": self.gps_lon} if self.gps_lat else None,
            "camera": f"{self.camera_make} {self.camera_model}".strip() if self.camera_make else None,
        }


# File type detection by extension
FILE_TYPE_MAP = {
    # Images
    ".jpg": "image", ".jpeg": "image", ".png": "image", ".tiff": "image",
    ".tif": "image", ".heic": "image", ".heif": "image", ".webp": "image",
    ".gif": "image", ".bmp": "image", ".raw": "image", ".cr2": "image",
    ".nef": "image", ".arw": "image", ".dng": "image",

    # Documents
    ".pdf": "pdf",
    ".doc": "office", ".docx": "office",
    ".xls": "office", ".xlsx": "office",
    ".ppt": "office", ".pptx": "office",
    ".odt": "office", ".ods": "office", ".odp": "office",

    # Videos
    ".mp4": "video", ".mov": "video", ".avi": "video",
    ".mkv": "video", ".wmv": "video", ".flv": "video",
}


def get_file_type(url: str) -> Optional[str]:
    """Determine file type from URL extension."""
    try:
        path = url.split("?")[0].lower()
        for ext, ftype in FILE_TYPE_MAP.items():
            if path.endswith(ext):
                return ftype
    except Exception:
        pass
    return None


def extract_metadata(data: bytes, url: str, mime_type: Optional[str] = None) -> MetadataResult:
    """
    Extract metadata from file bytes.

    Automatically detects file type and uses appropriate extractor.
    """
    file_type = get_file_type(url)

    if not file_type and mime_type:
        if "image" in mime_type:
            file_type = "image"
        elif "pdf" in mime_type:
            file_type = "pdf"
        elif "word" in mime_type or "document" in mime_type:
            file_type = "office"
        elif "video" in mime_type:
            file_type = "video"

    if not file_type:
        return MetadataResult(
            url=url,
            file_type="unknown",
            success=False,
            error="Could not determine file type"
        )

    # Route to appropriate extractor
    if file_type == "image":
        return extract_exif(data, url)
    elif file_type == "pdf":
        return extract_pdf_meta(data, url)
    elif file_type == "office":
        return extract_office_meta(data, url)
    elif file_type == "video":
        return extract_video_meta(data, url)

    return MetadataResult(
        url=url,
        file_type=file_type,
        success=False,
        error=f"No extractor for file type: {file_type}"
    )


def extract_exif(data: bytes, url: str) -> MetadataResult:
    """
    Extract EXIF metadata from images.

    Tries multiple methods:
    1. exiftool (most comprehensive)
    2. Pillow (fallback)
    3. piexif (JPEG only fallback)
    """
    result = MetadataResult(url=url, file_type="image", success=False)

    # Try exiftool first (most comprehensive)
    try:
        meta = _extract_with_exiftool(data)
        if meta:
            result.metadata = meta
            result.success = True
            _populate_common_fields(result, meta)
            return result
    except Exception as e:
        logger.debug(f"exiftool failed: {e}")

    # Fallback to Pillow
    try:
        meta = _extract_with_pillow(data)
        if meta:
            result.metadata = meta
            result.success = True
            _populate_common_fields(result, meta)
            return result
    except Exception as e:
        logger.debug(f"Pillow failed: {e}")

    # Fallback to piexif
    try:
        meta = _extract_with_piexif(data)
        if meta:
            result.metadata = meta
            result.success = True
            _populate_common_fields(result, meta)
            return result
    except Exception as e:
        logger.debug(f"piexif failed: {e}")

    result.error = "All EXIF extractors failed"
    return result


def _extract_with_exiftool(data: bytes) -> Optional[Dict]:
    """Extract using exiftool command-line tool."""
    try:
        # Write to temp file and run exiftool
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=".tmp") as f:
            f.write(data)
            temp_path = f.name

        try:
            result = subprocess.run(
                ["exiftool", "-json", "-a", "-g", temp_path],
                capture_output=True,
                text=True,
                timeout=30
            )

            if result.returncode == 0 and result.stdout:
                parsed = json.loads(result.stdout)
                if parsed and isinstance(parsed, list):
                    return parsed[0]
        finally:
            Path(temp_path).unlink(missing_ok=True)
    except FileNotFoundError:
        logger.debug("exiftool not installed")
    except Exception as e:
        logger.debug(f"exiftool error: {e}")

    return None


def _extract_with_pillow(data: bytes) -> Optional[Dict]:
    """Extract using Pillow library."""
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS, GPSTAGS

        img = Image.open(io.BytesIO(data))
        exif_data = img._getexif()

        if not exif_data:
            return None

        result = {}
        for tag_id, value in exif_data.items():
            tag = TAGS.get(tag_id, tag_id)

            # Handle GPS info specially
            if tag == "GPSInfo":
                gps_data = {}
                for gps_tag_id, gps_value in value.items():
                    gps_tag = GPSTAGS.get(gps_tag_id, gps_tag_id)
                    gps_data[gps_tag] = gps_value
                result["GPSInfo"] = gps_data
            else:
                # Convert bytes to string if needed
                if isinstance(value, bytes):
                    try:
                        value = value.decode("utf-8", errors="ignore")
                    except Exception:
                        value = str(value)
                result[tag] = value

        return result
    except ImportError:
        logger.debug("Pillow not installed")
    except Exception as e:
        logger.debug(f"Pillow error: {e}")

    return None


def _extract_with_piexif(data: bytes) -> Optional[Dict]:
    """Extract using piexif library (JPEG only)."""
    try:
        import piexif

        exif_dict = piexif.load(data)
        result = {}

        # Flatten the nested structure
        for ifd_name in ("0th", "Exif", "GPS", "1st"):
            ifd = exif_dict.get(ifd_name, {})
            for tag, value in ifd.items():
                tag_name = piexif.TAGS.get(ifd_name, {}).get(tag, {}).get("name", str(tag))

                # Decode bytes
                if isinstance(value, bytes):
                    try:
                        value = value.decode("utf-8", errors="ignore").strip("\x00")
                    except Exception:
                        value = str(value)

                result[tag_name] = value

        return result if result else None
    except ImportError:
        logger.debug("piexif not installed")
    except Exception as e:
        logger.debug(f"piexif error: {e}")

    return None


def extract_pdf_meta(data: bytes, url: str) -> MetadataResult:
    """Extract metadata from PDF files."""
    result = MetadataResult(url=url, file_type="pdf", success=False)

    try:
        import pypdf

        reader = pypdf.PdfReader(io.BytesIO(data))
        meta = {}

        if reader.metadata:
            for key, value in reader.metadata.items():
                # Clean up key name (remove leading /)
                clean_key = key.lstrip("/")
                meta[clean_key] = str(value) if value else None

        meta["PageCount"] = len(reader.pages)

        result.metadata = meta
        result.success = True
        _populate_common_fields(result, meta)

    except ImportError:
        result.error = "pypdf not installed"
    except Exception as e:
        result.error = f"PDF extraction failed: {e}"

    return result


def extract_office_meta(data: bytes, url: str) -> MetadataResult:
    """Extract metadata from Office documents (DOCX, XLSX, PPTX)."""
    result = MetadataResult(url=url, file_type="office", success=False)

    # Determine specific type
    path_lower = url.lower()

    if ".docx" in path_lower:
        return _extract_docx_meta(data, url)
    elif ".xlsx" in path_lower:
        return _extract_xlsx_meta(data, url)
    elif ".pptx" in path_lower:
        return _extract_pptx_meta(data, url)
    elif ".doc" in path_lower or ".xls" in path_lower or ".ppt" in path_lower:
        # Try OLE extraction for legacy formats
        return _extract_ole_meta(data, url)

    result.error = "Unsupported Office format"
    return result


def _extract_docx_meta(data: bytes, url: str) -> MetadataResult:
    """Extract from DOCX files."""
    result = MetadataResult(url=url, file_type="docx", success=False)

    try:
        import docx

        doc = docx.Document(io.BytesIO(data))
        core_props = doc.core_properties

        meta = {
            "Author": core_props.author,
            "Creator": core_props.author,
            "Title": core_props.title,
            "Subject": core_props.subject,
            "Keywords": core_props.keywords,
            "LastModifiedBy": core_props.last_modified_by,
            "Created": str(core_props.created) if core_props.created else None,
            "Modified": str(core_props.modified) if core_props.modified else None,
            "Category": core_props.category,
            "Comments": core_props.comments,
            "Version": core_props.version,
            "Revision": core_props.revision,
        }

        # Remove None values
        meta = {k: v for k, v in meta.items() if v}

        result.metadata = meta
        result.success = True
        _populate_common_fields(result, meta)

    except ImportError:
        result.error = "python-docx not installed"
    except Exception as e:
        result.error = f"DOCX extraction failed: {e}"

    return result


def _extract_xlsx_meta(data: bytes, url: str) -> MetadataResult:
    """Extract from XLSX files."""
    result = MetadataResult(url=url, file_type="xlsx", success=False)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        props = wb.properties

        meta = {
            "Author": props.creator,
            "Creator": props.creator,
            "Title": props.title,
            "Subject": props.subject,
            "Keywords": props.keywords,
            "LastModifiedBy": props.lastModifiedBy,
            "Created": str(props.created) if props.created else None,
            "Modified": str(props.modified) if props.modified else None,
            "Category": props.category,
            "Company": props.company,
            "Description": props.description,
            "SheetCount": len(wb.sheetnames),
            "SheetNames": wb.sheetnames,
        }

        meta = {k: v for k, v in meta.items() if v}

        result.metadata = meta
        result.success = True
        _populate_common_fields(result, meta)

    except ImportError:
        result.error = "openpyxl not installed"
    except Exception as e:
        result.error = f"XLSX extraction failed: {e}"

    return result


def _extract_pptx_meta(data: bytes, url: str) -> MetadataResult:
    """Extract from PPTX files."""
    result = MetadataResult(url=url, file_type="pptx", success=False)

    try:
        from pptx import Presentation

        prs = Presentation(io.BytesIO(data))
        props = prs.core_properties

        meta = {
            "Author": props.author,
            "Creator": props.author,
            "Title": props.title,
            "Subject": props.subject,
            "Keywords": props.keywords,
            "LastModifiedBy": props.last_modified_by,
            "Created": str(props.created) if props.created else None,
            "Modified": str(props.modified) if props.modified else None,
            "Category": props.category,
            "Comments": props.comments,
            "SlideCount": len(prs.slides),
        }

        meta = {k: v for k, v in meta.items() if v}

        result.metadata = meta
        result.success = True
        _populate_common_fields(result, meta)

    except ImportError:
        result.error = "python-pptx not installed"
    except Exception as e:
        result.error = f"PPTX extraction failed: {e}"

    return result


def _extract_ole_meta(data: bytes, url: str) -> MetadataResult:
    """Extract from legacy Office formats using OLE."""
    result = MetadataResult(url=url, file_type="ole", success=False)

    try:
        import olefile

        ole = olefile.OleFileIO(io.BytesIO(data))
        meta = ole.get_metadata()

        result_meta = {
            "Author": meta.author,
            "Title": meta.title,
            "Subject": meta.subject,
            "Keywords": meta.keywords,
            "LastSavedBy": meta.last_saved_by,
            "CreatingApplication": meta.creating_application,
            "Company": meta.company,
            "Created": str(meta.create_time) if meta.create_time else None,
            "Modified": str(meta.last_saved_time) if meta.last_saved_time else None,
        }

        result_meta = {k: v for k, v in result_meta.items() if v}

        result.metadata = result_meta
        result.success = True
        _populate_common_fields(result, result_meta)

    except ImportError:
        result.error = "olefile not installed"
    except Exception as e:
        result.error = f"OLE extraction failed: {e}"

    return result


def extract_video_meta(data: bytes, url: str) -> MetadataResult:
    """Extract metadata from video files using exiftool."""
    result = MetadataResult(url=url, file_type="video", success=False)

    # Video metadata is best extracted with exiftool
    try:
        meta = _extract_with_exiftool(data)
        if meta:
            result.metadata = meta
            result.success = True
            _populate_common_fields(result, meta)
            return result
    except Exception as e:
        result.error = f"Video extraction failed: {e}"

    return result


def _populate_common_fields(result: MetadataResult, meta: Dict):
    """Populate common investigation-relevant fields from metadata."""

    # Author/Creator
    for key in ("Author", "Creator", "Artist", "By-line", "Owner", "XPAuthor"):
        if key in meta and meta[key]:
            result.author = str(meta[key])
            break

    for key in ("Creator", "CreatorTool", "CreatingApplication"):
        if key in meta and meta[key]:
            result.creator = str(meta[key])
            break

    # Company/Organization
    for key in ("Company", "Organization", "XPComment"):
        if key in meta and meta[key]:
            result.company = str(meta[key])
            break

    # Software
    for key in ("Software", "CreatorTool", "Producer", "CreatingApplication"):
        if key in meta and meta[key]:
            result.software = str(meta[key])
            break

    # Dates
    # DateTimeOriginal is specifically when the photo was taken (EXIF)
    if "DateTimeOriginal" in meta and meta["DateTimeOriginal"]:
        result.original_date = str(meta["DateTimeOriginal"])

    # Create date (file creation, not photo capture)
    for key in ("CreateDate", "Created", "CreationDate"):
        if key in meta and meta[key]:
            result.create_date = str(meta[key])
            break
    # Fallback to DateTimeOriginal if no create_date
    if not result.create_date and result.original_date:
        result.create_date = result.original_date

    for key in ("ModifyDate", "Modified", "ModDate", "DateTime"):
        if key in meta and meta[key]:
            result.modify_date = str(meta[key])
            break

    # GPS
    if "GPSInfo" in meta:
        gps = meta["GPSInfo"]
        try:
            result.gps_lat = _convert_gps(gps.get("GPSLatitude"), gps.get("GPSLatitudeRef"))
            result.gps_lon = _convert_gps(gps.get("GPSLongitude"), gps.get("GPSLongitudeRef"))
        except Exception:
            pass
    elif "GPSLatitude" in meta:
        try:
            result.gps_lat = float(meta["GPSLatitude"])
            result.gps_lon = float(meta.get("GPSLongitude", 0))
        except Exception:
            pass

    # Camera
    result.camera_make = meta.get("Make") or meta.get("CameraMake")
    result.camera_model = meta.get("Model") or meta.get("CameraModel")


def _convert_gps(coord, ref) -> Optional[float]:
    """Convert GPS coordinates to decimal degrees."""
    if not coord:
        return None

    try:
        if isinstance(coord, (list, tuple)) and len(coord) >= 3:
            # DMS format: (degrees, minutes, seconds)
            d = float(coord[0])
            m = float(coord[1])
            s = float(coord[2])
            decimal = d + m/60 + s/3600
        else:
            decimal = float(coord)

        if ref in ("S", "W"):
            decimal = -decimal

        return decimal
    except Exception:
        return None
